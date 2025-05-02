from http.client import responses
import pandas as pd
import json
import requests
import argparse
from datetime import datetime, timedelta
import openpyxl
from dateutil import relativedelta
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Font


def parse_args():
    parser = argparse.ArgumentParser()

    parser.add_argument("csv_file", help="CSV-Datei mit Fahrzeugdaten")
    parser.add_argument("-k", "--keys", nargs="*", help="Zusätzliche Spalten")
    parser.add_argument("-c", "--colored", action="store_true", help="aktiviere farbliche Markierung")

    return parser.parse_args()


def send_csv_to_server(path: str) -> list[dict]:
    """
    Sends a POST-Message to the Server-API and returns the enriched data as a list of dictionaries
    args: path (str) the path to the csv_file you want to send
    returns: list of dictionaries
    """
    url = "http://localhost:8000/upload.csv/"
    with open(path, "rb") as f:
        files = {"file": (path, f, "text/csv")}
        response = requests.post(url, files=files)

        return response.json()

def calculate_hu_age(hu: str) -> int:
   """
   Expects the date of the hu (iso formatted), calculates the age (in months) of the hu with the current date and
   returns it
   args: hu (str)
   returns age in months (int)
   """
   hu_datelist =  hu.split("-")
   hu_date = datetime(int(hu_datelist[0]), int(hu_datelist[1]), int(hu_datelist[2]))
   today = datetime.now()
   delta = relativedelta.relativedelta(today, hu_date)

   return delta.years * 12 + delta.months



def create_excel(additional_columns: list, tinted: bool, csv_data: list[dict]):
    """
    Expects the additional columns which are given in the command by -k, a tinted value (also given in the
    command (-c), csv_data as a list of dictionaies and creates an excel-file of the csv_data
    args: additional_columns (list), tinted (bool), csv_data (list of dictionaries)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "vehicles"
    column_names = ["rnr"]

    if additional_columns:
        column_names.extend(additional_columns)
    ws.append(column_names)

    # color codes for the hu
    red_fill = PatternFill(start_color="FFb30000", fill_type="solid")
    orange_fill = PatternFill(start_color="FFFFA500", fill_type="solid")
    green_fill = PatternFill(start_color="FF007500", fill_type="solid")

    for idx, row in enumerate(csv_data, start=2):
        row_data = [row.get("rnr")] + [row.get(col) for col in additional_columns]
        ws.append(row_data)

        if tinted and "hu" in row:
            age_in_months = calculate_hu_age(row["hu"])
            if age_in_months <= 3:
                fill = green_fill
            elif age_in_months <= 12:
                fill = orange_fill
            else:
                fill = red_fill
            for col in range(1, len(row_data) + 1):
                ws.cell(row=idx, column=col).fill = fill

        if "labelIds" in column_names and "colorCode" in row and row["colorCode"]:
            colorcode = row["colorCode"].strip("#")
            color = Font(color=colorcode)
            try:
                col_index = column_names.index("labelIds") +1
                ws.cell(row=idx, column=col_index).font = color
            except:
                pass

    wb.save(f"vehicles_{datetime.now().date().isoformat()}.xlsx")


if __name__ == "__main__":
    args = parse_args()

    print("CSV-Datei wird verarbeitet...")
    response = send_csv_to_server(args.csv_file)

    df = pd.DataFrame(response)
    df = df.sort_values(by="gruppe")
    sorted_data = json.loads(df.to_json(orient="records", indent=2))
    

    create_excel(list(args.keys), args.colored, sorted_data)
